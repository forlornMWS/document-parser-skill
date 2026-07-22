#!/usr/bin/env python3
"""
Excel 模板格式解析器
将 xlsx 文件解析为结构化 JSON，供 Agent 生成 PDF 时参考样式。

用法:
    python parse_excel.py <input.xlsx> [output.json]
    python parse_excel.py template.xlsx                    # 输出到 template_format.json
    python parse_excel.py template.xlsx -                  # 输出到 stdout
    python parse_excel.py template.xlsx --pretty           # 美化输出(带缩进)
    python parse_excel.py template.xlsx --skip-empty       # 跳过空单元格
    python parse_excel.py template.xlsx --sheet "Sheet1,Sheet3"  # 只解析指定 sheet
    python parse_excel.py template.xlsx --exclude "Sheet2"       # 排除指定 sheet
    python parse_excel.py template.xlsx --list-sheets            # 列出所有 sheet
"""

import sys
import json
import argparse
from datetime import datetime, date, time
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 常量映射 ──────────────────────────────────────────────

PAPER_SIZES = {
    1: "Letter", 2: "Letter", 3: "Tabloid", 4: "Ledger",
    5: "Legal", 6: "Statement", 7: "Executive", 8: "A3",
    9: "A4", 10: "A4", 11: "A5", 12: "B4", 13: "B5",
}

BORDER_STYLES = {
    None: "none",
    "thin": "thin",
    "medium": "medium",
    "dashed": "dashed",
    "dotted": "dotted",
    "thick": "thick",
    "double": "double",
    "hair": "hair",
    "mediumDashed": "mediumDashed",
    "dashDot": "dashDot",
    "mediumDashDot": "mediumDashDot",
    "dashDotDot": "dashDotDot",
    "mediumDashDotDot": "mediumDashDotDot",
    "slantDashDot": "slantDashDot",
}


def color_to_dict(color):
    """将 openpyxl Color 对象转为字典"""
    if color is None:
        return None
    result = {}
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        result["rgb"] = str(color.rgb)
    elif color.type == "theme":
        result["theme"] = color.theme
        if color.tint and color.tint != 0:
            result["tint"] = round(color.tint, 4)
    elif color.type == "indexed":
        result["indexed"] = color.indexed
    return result if result else None


def parse_font(font):
    """解析字体样式"""
    if font is None:
        return None
    result = {}
    if font.name:
        result["name"] = font.name
    if font.size:
        result["size"] = font.size
    if font.bold:
        result["bold"] = True
    if font.italic:
        result["italic"] = True
    if font.underline and font.underline != "none":
        result["underline"] = font.underline
    if font.strikethrough:
        result["strikethrough"] = True
    color = color_to_dict(font.color)
    if color:
        result["color"] = color
    return result if result else None


def parse_alignment(alignment):
    """解析对齐方式"""
    if alignment is None:
        return None
    result = {}
    # 水平对齐：仅过滤 "general"（Excel 默认值）
    if alignment.horizontal and alignment.horizontal != "general":
        result["horizontal"] = alignment.horizontal
    # 垂直对齐：始终输出（bottom 虽然是默认值，但显式设置时应保留）
    if alignment.vertical:
        result["vertical"] = alignment.vertical
    if alignment.wrap_text:
        result["wrap_text"] = True
    if alignment.text_rotation and alignment.text_rotation != 0:
        result["text_rotation"] = alignment.text_rotation
    if alignment.indent and alignment.indent != 0:
        result["indent"] = alignment.indent
    return result if result else None


def parse_fill(fill):
    """解析背景填充"""
    if fill is None:
        return None
    result = {}
    # openpyxl 使用 patternType 而非 pattern_type
    pattern_type = getattr(fill, 'patternType', None) or getattr(fill, 'pattern_type', None)
    if pattern_type and pattern_type != "none":
        result["pattern_type"] = pattern_type
    fg = color_to_dict(fill.fgColor)
    if fg:
        result["fg_color"] = fg
    bg = color_to_dict(fill.bgColor)
    if bg:
        result["bg_color"] = bg
    return result if result else None


def parse_border_side(side):
    """解析单边边框"""
    if side is None or side.style is None:
        return None
    result = {"style": BORDER_STYLES.get(side.style, str(side.style))}
    color = color_to_dict(side.color)
    if color:
        result["color"] = color
    return result


def parse_border(border):
    """解析边框"""
    if border is None:
        return None
    result = {}
    for side_name in ["top", "bottom", "left", "right"]:
        side = getattr(border, side_name, None)
        parsed = parse_border_side(side)
        if parsed:
            result[side_name] = parsed
    # 对角线
    if border.diagonal and border.diagonal.style:
        result["diagonal"] = parse_border_side(border.diagonal)
        if border.diagonal_down:
            result["diagonal_down"] = True
        if border.diagonal_up:
            result["diagonal_up"] = True
    return result if result else None


def parse_style(cell):
    """解析单元格完整样式"""
    style = {}
    font = parse_font(cell.font)
    if font:
        style["font"] = font
    alignment = parse_alignment(cell.alignment)
    if alignment:
        style["alignment"] = alignment
    fill = parse_fill(cell.fill)
    if fill:
        style["fill"] = fill
    border = parse_border(cell.border)
    if border:
        style["border"] = border
    return style if style else None


def detect_data_type(cell):
    """检测单元格数据类型"""
    if cell.value is None:
        return "empty"
    if isinstance(cell.value, bool):
        return "boolean"
    if isinstance(cell.value, (int, float)):
        return "number"
    if isinstance(cell.value, datetime):
        return "datetime"
    if isinstance(cell.value, date):
        return "date"
    if isinstance(cell.value, time):
        return "time"
    if isinstance(cell.value, str):
        if cell.value.startswith("="):
            return "formula"
        return "string"
    return "string"


def format_cell_value(cell):
    """获取单元格的格式化显示值"""
    if cell.value is None:
        return None
    # 如果是日期类型，返回 ISO 格式
    if isinstance(cell.value, datetime):
        return cell.value.isoformat()
    if isinstance(cell.value, date):
        return cell.value.isoformat()
    if isinstance(cell.value, time):
        return cell.value.isoformat()
    return cell.value


def parse_sheet_view(ws):
    """解析工作表视图设置（缩放、网格线、冻结窗格等）"""
    result = {}
    
    # 缩放
    if ws.sheet_view.zoomScale is not None:
        result['zoom_scale'] = ws.sheet_view.zoomScale
    if ws.sheet_view.zoomScaleNormal is not None:
        result['zoom_scale_normal'] = ws.sheet_view.zoomScaleNormal
    if ws.sheet_view.zoomScalePageLayoutView is not None:
        result['zoom_scale_page_layout'] = ws.sheet_view.zoomScalePageLayoutView
    
    # 网格线
    if ws.sheet_view.showGridLines is not None:
        result['show_grid_lines'] = ws.sheet_view.showGridLines
    
    # 冻结窗格
    if ws.freeze_panes:
        result['freeze_panes'] = ws.freeze_panes
    
    # 其他视图属性
    if ws.sheet_view.showFormulas:
        result['show_formulas'] = True
    if ws.sheet_view.showRowColHeaders is False:
        result['show_headers'] = False
    if ws.sheet_view.showZeros is False:
        result['show_zeros'] = False
    if ws.sheet_view.rightToLeft:
        result['right_to_left'] = True
    if ws.sheet_view.tabSelected:
        result['tab_selected'] = True
    
    return result if result else None


def parse_page_setup(ws):
    """解析页面设置"""
    ps = ws.page_setup
    pm = ws.page_margins

    paper_code = ps.paperSize
    paper_name = PAPER_SIZES.get(paper_code, "A4") if paper_code else "A4"

    result = {
        "paper_size": paper_name,
        "orientation": ps.orientation if ps.orientation else "portrait",
        "margins": {
            "top": round(pm.top * 25.4, 2),       # 英寸 → mm
            "bottom": round(pm.bottom * 25.4, 2),
            "left": round(pm.left * 25.4, 2),
            "right": round(pm.right * 25.4, 2),
            "header": round(pm.header * 25.4, 2),
            "footer": round(pm.footer * 25.4, 2),
        }
    }

    # 打印区域
    if ws.print_area:
        result["print_area"] = ws.print_area

    # 缩放
    try:
        if ps.fitToPage:
            result["fit_to_page"] = True
    except (AttributeError, TypeError):
        pass
    try:
        if ps.scale:
            result["scale"] = ps.scale
    except (AttributeError, TypeError):
        pass

    # 打印标题行/列
    if ws.print_title_rows:
        result["print_title_rows"] = ws.print_title_rows
    if ws.print_title_cols:
        result["print_title_cols"] = ws.print_title_cols

    return result


def parse_columns(ws):
    """解析列定义"""
    columns = []
    for col_letter, dim in ws.column_dimensions.items():
        if len(col_letter) > 2:  # 跳过无效标识
            continue
        # 使用 openpyxl 工具函数将列字母转为 1-based 索引
        col_idx = column_index_from_string(col_letter) - 1  # 转为 0-based

        col_info = {
            "index": col_idx,
            "letter": col_letter,
            "width": dim.width if dim.width else 8.43,
            "width_px": round((dim.width if dim.width else 8.43) * 7.5, 1),
        }
        if dim.hidden:
            col_info["hidden"] = True
        columns.append(col_info)

    # 按列索引排序
    columns.sort(key=lambda c: c["index"])
    return columns


def parse_rows(ws):
    """解析行定义"""
    rows = []
    for row_num, dim in ws.row_dimensions.items():
        row_info = {
            "index": row_num - 1,
            "number": row_num,
            "height": dim.height if dim.height else 15,
            "height_px": round((dim.height if dim.height else 15) * 1.333, 1),
        }
        if dim.hidden:
            row_info["hidden"] = True
        if dim.customHeight:
            row_info["custom_height"] = True
        rows.append(row_info)

    rows.sort(key=lambda r: r["index"])
    return rows


def parse_merged_cells(ws):
    """解析合并单元格"""
    merged = []
    for merge_range in ws.merged_cells.ranges:
        range_str = str(merge_range)
        merged.append({
            "range": range_str,
            "start_row": merge_range.min_row - 1,
            "start_col": merge_range.min_col - 1,
            "end_row": merge_range.max_row - 1,
            "end_col": merge_range.max_col - 1,
            "rowspan": merge_range.max_row - merge_range.min_row + 1,
            "colspan": merge_range.max_col - merge_range.min_col + 1,
        })
    return merged


def parse_hyperlinks(ws):
    """解析工作表中的超链接"""
    links = []
    # 优先使用公开属性 hyperlinks（openpyxl 较新版本）
    hyperlinks = getattr(ws, 'hyperlinks', None)
    if hyperlinks is None:
        # 旧版本回退到私有属性
        hyperlinks = getattr(ws, '_hyperlinks', [])
    for link in hyperlinks:
        link_data = {
            "ref": str(link.ref),
            "target": link.target or "",
        }
        if link.display:
            link_data["display"] = link.display
        if link.tooltip:
            link_data["tooltip"] = link.tooltip
        links.append(link_data)
    return links


def parse_cells(ws_formula, ws_data, skip_empty=False, hyperlinks=None):
    """解析所有单元格
    
    Args:
        ws_formula: data_only=False 的工作表，用于获取公式
        ws_data: data_only=True 的工作表，用于获取计算值
        skip_empty: 是否跳过空单元格
        hyperlinks: 超链接列表
    """
    cells = []
    # 构建超链接索引
    link_map = {}
    if hyperlinks:
        for link in hyperlinks:
            link_map[link["ref"]] = link
    
    for row_formula, row_data in zip(ws_formula.iter_rows(), ws_data.iter_rows()):
        for cell_formula, cell_data in zip(row_formula, row_data):
            # 跳过完全空白且无样式的单元格
            if skip_empty and cell_formula.value is None and not cell_formula.has_style:
                continue

            cell_info = {
                "address": cell_formula.coordinate,
                "row": cell_formula.row - 1,
                "col": cell_formula.column - 1,
            }

            # 值和数据类型
            data_type = detect_data_type(cell_formula)
            cell_info["data_type"] = data_type

            if data_type == "formula":
                # 公式单元格：同时输出公式和计算值
                cell_info["formula"] = cell_formula.value
                cell_info["value"] = format_cell_value(cell_data)
            elif data_type != "empty":
                cell_info["value"] = format_cell_value(cell_formula)

            # 数字格式
            if cell_formula.number_format and cell_formula.number_format != "General":
                cell_info["number_format"] = cell_formula.number_format

            # 样式
            style = parse_style(cell_formula)
            if style:
                cell_info["style"] = style

            # 超链接
            if cell_formula.coordinate in link_map:
                cell_info["hyperlink"] = link_map[cell_formula.coordinate]["target"]

            cells.append(cell_info)

    return cells


def parse_workbook(filepath, skip_empty=False, sheet_filter=None, sheet_exclude=None):
    """解析整个工作簿
    
    Args:
        filepath: xlsx 文件路径
        skip_empty: 是否跳过空单元格
        sheet_filter: 指定要解析的 sheet 名称列表（None 表示全部）
        sheet_exclude: 要排除的 sheet 名称列表
    """
    # 加载两次工作簿：一次获取公式，一次获取计算值
    wb_formula = load_workbook(filepath, data_only=False)
    wb_data = load_workbook(filepath, data_only=True)

    try:
        result = {
            "metadata": {
                "source_file": Path(filepath).name,
                "parsed_at": datetime.now().isoformat(),
                "parser_version": "1.0.0",
            },
            "sheets": []
        }

        # 确定要解析的 sheet 列表
        target_sheets = wb_formula.sheetnames
        if sheet_filter:
            target_sheets = [s for s in target_sheets if s in sheet_filter]
        if sheet_exclude:
            target_sheets = [s for s in target_sheets if s not in sheet_exclude]

        for ws_name in target_sheets:
            ws_formula = wb_formula[ws_name]
            ws_data = wb_data[ws_name]
            
            # 解析超链接
            hyperlinks = parse_hyperlinks(ws_formula)
            
            sheet_data = {
                "name": ws_name,
                "dimensions": ws_formula.dimensions,
                "max_row": ws_formula.max_row,
                "max_column": ws_formula.max_column,
                "page_setup": parse_page_setup(ws_formula),
                "columns": parse_columns(ws_formula),
                "rows": parse_rows(ws_formula),
                "merged_cells": parse_merged_cells(ws_formula),
                "cells": parse_cells(ws_formula, ws_data, skip_empty=skip_empty, hyperlinks=hyperlinks),
            }
            
            # 工作表视图设置
            sheet_view = parse_sheet_view(ws_formula)
            if sheet_view:
                sheet_data["sheet_view"] = sheet_view
            
            result["sheets"].append(sheet_data)

        return result
    finally:
        wb_formula.close()
        wb_data.close()


def main():
    parser = argparse.ArgumentParser(
        description="Excel 模板格式解析器 - 将 xlsx 转为结构化 JSON"
    )
    parser.add_argument("input", help="输入的 xlsx 文件路径")
    parser.add_argument("output", nargs="?", default=None,
                        help="输出 JSON 文件路径 (默认: <input>_format.json, 用 - 输出到 stdout)")
    parser.add_argument("--pretty", action="store_true",
                        help="美化输出(带缩进，默认紧凑)")
    parser.add_argument("--skip-empty", action="store_true",
                        help="跳过空单元格(无值且无样式)")
    parser.add_argument("--schema", action="store_true",
                        help="同时输出 JSON Schema 文件路径提示")
    parser.add_argument("--sheet", type=str, default=None,
                        help="指定要解析的 sheet，多个用逗号分隔 (如: Sheet1,Sheet3)")
    parser.add_argument("--exclude", type=str, default=None,
                        help="排除指定 sheet，多个用逗号分隔 (如: Sheet2,Sheet4)")
    parser.add_argument("--list-sheets", action="store_true",
                        help="列出所有 sheet 名称后退出")

    args = parser.parse_args()

    # 检查输入文件
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    # 列出 sheet 模式
    if args.list_sheets:
        wb = load_workbook(args.input, read_only=True)
        print(f"工作簿包含 {len(wb.sheetnames)} 个 sheet:")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {name}")
        wb.close()
        sys.exit(0)

    # 解析 sheet 过滤参数
    sheet_filter = None
    sheet_exclude = None
    if args.sheet:
        sheet_filter = [s.strip() for s in args.sheet.split(",")]
    if args.exclude:
        sheet_exclude = [s.strip() for s in args.exclude.split(",")]

    # 解析
    print(f"Parsing: {args.input}", file=sys.stderr)
    data = parse_workbook(args.input, skip_empty=args.skip_empty, 
                         sheet_filter=sheet_filter, sheet_exclude=sheet_exclude)

    # 统计
    total_cells = sum(len(s["cells"]) for s in data["sheets"])
    total_merged = sum(len(s["merged_cells"]) for s in data["sheets"])
    print(f"Sheets: {len(data['sheets'])}, Cells: {total_cells}, Merged: {total_merged}", file=sys.stderr)

    # 输出
    indent = 2 if args.pretty else None
    json_str = json.dumps(data, indent=indent, ensure_ascii=False, default=str)

    if args.output == "-":
        sys.stdout.write(json_str)
        sys.stdout.write("\n")
    else:
        output_path = args.output or str(input_path.with_stem(input_path.stem + "_format"))
        if not output_path.endswith(".json"):
            output_path += ".json"
        Path(output_path).write_text(json_str, encoding="utf-8")
        print(f"Output: {output_path}", file=sys.stderr)

    if args.schema:
        print("\nJSON Schema: excel-format-schema.json", file=sys.stderr)


if __name__ == "__main__":
    main()
